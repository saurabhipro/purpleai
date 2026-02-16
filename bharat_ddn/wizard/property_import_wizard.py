from odoo import models, fields, api
from odoo.exceptions import ValidationError
import base64
import csv
from io import StringIO, BytesIO
import os
import tempfile
import openpyxl
from odoo.exceptions import UserError
from odoo.tools import _
import logging

_logger = logging.getLogger(__name__)

class PropertyImportWizard(models.TransientModel):
    _name = 'property.import.wizard'
    _description = 'Wizard to import property data'

    data_file = fields.Binary('CSV or Excel File', required=True)
    filename = fields.Char('Filename')
    sheet_name = fields.Char('Sheet Name', help="Name of the sheet to import. Defaults to the first sheet if empty.")
    available_sheets = fields.Text('Available Sheets', readonly=True)
    import_type = fields.Selection([
        ('property_id', 'Property ID Data'),
        ('upic', 'UPIC Data')
    ], string='Import Type', default='property_id', required=True)
    import_limit_option = fields.Selection([
        ('all', 'All Records'),
        ('limit', 'Limit Records')
    ], string='Import Option', default='all', required=True)
    limit_count = fields.Integer('No of Records', default=10)
    error_file = fields.Binary('Error File', readonly=True)
    error_filename = fields.Char('Error Filename', readonly=True)

    def get_headers_map(self):
        if self.import_type == 'upic':
            return {
                'company': 'company_id',
                'upicno': 'upic_no',
                'unit_no': 'unit_no',
                'zone': 'zone_id',
                'ward': 'ward_id',
                'colony': 'colony_id',
            }
        else:
            return {
                'zone': 'zone_id',
                'ward': 'ward_id',
                'property id': 'property_id',
                'owner name': 'owner_name',
                'address': 'address_line_1',
                'mobile no': 'mobile_no',
                'colony': 'colony_id',
            }

    @api.onchange('data_file')
    def _onchange_data_file(self):
        if self.data_file and self.filename:
            ext = os.path.splitext(self.filename)[-1].lower()
            if ext in ('.xls', '.xlsx'):
                try:
                    file_content = base64.b64decode(self.data_file)
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                        tmp.write(file_content)
                        tmp.seek(0)
                        wb = openpyxl.load_workbook(tmp.name, read_only=True)
                        formatted_sheets = [] 
                        for s in wb.sheetnames:
                             formatted_sheets.append(s)
                        
                        self.available_sheets = ", ".join(formatted_sheets)
                        if formatted_sheets:
                            self.sheet_name = formatted_sheets[0]
                        
                        os.unlink(tmp.name)
                except Exception as e:
                    _logger.warning(f"Could not read excel sheets: {e}")
                    self.available_sheets = "Could not read sheets."

    def action_import(self):
        if not self.data_file:
            raise ValidationError('Please upload a CSV or Excel file.')
        if not self.filename:
            raise ValidationError('Filename is missing.')
        
        ext = os.path.splitext(self.filename)[-1].lower()
        skipped_records = []
        batch_size = 100
        current_batch = []
        created_records = 0
        
        try:
            file_content = base64.b64decode(self.data_file)
            rows = []
            
            if ext == '.csv':
                # Simplified CSV support - assuming standard mapping slightly different from Excel for now specific to requester or just generic
                # But requirement focuses on Excel tabs.
                csvfile = StringIO(file_content.decode('utf-8'))
                reader = csv.DictReader(csvfile)
                # Map CSV headers loosely
                rows = list(reader)
                # TODO: Adapt CSV logic if needed, but focusing on Excel as per request
            elif ext in ('.xls', '.xlsx'):
                try:
                    import pandas as pd
                except ImportError:
                    pass # pandas optional if we use openpyxl directly
                
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(file_content)
                    tmp.seek(0)
                    wb = openpyxl.load_workbook(tmp.name, data_only=True)
                    
                    target_sheet = None
                    if self.sheet_name:
                        if self.sheet_name in wb.sheetnames:
                            target_sheet = wb[self.sheet_name]
                        else:
                            raise ValidationError(f"Sheet '{self.sheet_name}' not found. Available: {wb.sheetnames}")
                    else:
                        target_sheet = wb.active
                        
                    sheet = target_sheet
                    
                    # Header Validation
                    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                    if not headers:
                        raise ValidationError("Sheet is empty or missing headers.")
                    
                    # Normalize headers
                    headers_lower = [str(h).strip().lower() for h in headers if h]
                    header_map_config = self.get_headers_map()
                    
                    # Check missing
                    missing = []
                    col_index_map = {}
                    
                    found_headers_map = {}
                    for idx, h in enumerate(headers):
                        if not h: continue
                        h_clean = str(h).strip().lower()
                        found_headers_map[h_clean] = idx

                    for req_key in header_map_config.keys():
                        if req_key not in found_headers_map:
                            missing.append(req_key)
                        else:
                            col_index_map[req_key] = found_headers_map[req_key]
                    
                    if missing:
                        msg = f"Missing required columns in Excel: {', '.join(missing)}.\n"
                        msg += f"Expected columns: {', '.join(header_map_config.keys())}"
                        raise UserError(msg)

                    rows = []
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        # Extract data using the map
                        row_data = {}
                        has_data = False
                        for key, col_idx in col_index_map.items():
                            val = row[col_idx] if col_idx < len(row) else None
                            row_data[key] = val
                            if val: has_data = True
                        
                        if not has_data: continue # Skip empty rows

                        vals = {}
                        
                        # Process Zone
                        zone_name = row_data.get('zone')
                        if zone_name:
                            zone_id = self.env['ddn.zone'].search([('name', 'ilike', str(zone_name).strip())], limit=1)
                            if not zone_id:
                                # Dictionary to log error? Or skip?
                                # For now, let's skip/log or keep False
                                pass 
                            vals['zone_id'] = zone_id.id if zone_id else False

                        # Process Ward
                        ward_name = row_data.get('ward')
                        if ward_name:
                            ward_id = self.env['ddn.ward'].search([('name', 'ilike', str(ward_name).strip())], limit=1)
                            vals['ward_id'] = ward_id.id if ward_id else False

                        # Process Colony
                        colony_name = row_data.get('colony')
                        if colony_name:
                            colony_id = self.env['ddn.colony'].search([('name', 'ilike', str(colony_name).strip())], limit=1)
                            vals['colony_id'] = colony_id.id if colony_id else False


                        # Process Company (UPIC Only)
                        company_name = row_data.get('company')
                        if company_name:
                             company_id = self.env['res.company'].search([('name', 'ilike', str(company_name).strip())], limit=1)
                             vals['company_id'] = company_id.id if company_id else False

                        vals['property_id'] = str(row_data.get('property id')).strip() if row_data.get('property id') else False
                        vals['upic_no'] = str(row_data.get('upicno')).strip() if row_data.get('upicno') else False
                        vals['unit_no'] = str(row_data.get('unit_no')).strip() if row_data.get('unit_no') else False
                        vals['owner_name'] = row_data.get('owner name')
                        vals['address_line_1'] = row_data.get('address')
                        vals['mobile_no'] = row_data.get('mobile no')
                        
                        # Fundamental check
                        if self.import_type == 'property_id' and not vals.get('property_id'):
                             _logger.info(f"Row {row_idx}: Skipped due to missing Property ID")
                             continue
                        if self.import_type == 'upic' and not vals.get('upic_no'):
                             _logger.info(f"Row {row_idx}: Skipped due to missing UPIC No")
                             continue

                        rows.append(vals)
                        
                    os.unlink(tmp.name)

            else:
                raise ValidationError('Unsupported file type. Please upload a .csv, .xls, or .xlsx file.')
            
            if self.import_limit_option == 'limit' and self.limit_count > 0:
                 rows = rows[:self.limit_count]

            # Process rows and check for duplicates
            total_records = len(rows)
            processed_records = 0
            
            _logger.info(f"Starting import of {total_records} records")
            
            for row in rows:
                try:
                    existing_property = False
                    if self.import_type == 'property_id':
                        # Check if property with same Property ID exists
                        if row.get('property_id'):
                            existing_property = self.env['ddn.property.info'].search([
                                ('property_id', '=', row['property_id'])
                            ], limit=1)
                    elif self.import_type == 'upic':
                        # Check if property with same UPIC exists
                        if row.get('upic_no'):
                             existing_property = self.env['ddn.property.info'].search([
                                ('upic_no', '=', row['upic_no'])
                            ], limit=1)
                    
                    if not existing_property:
                        current_batch.append(row)
                        processed_records += 1
                        
                        # Process batch when it reaches batch_size
                        if len(current_batch) >= batch_size:
                            try:
                                # Create records
                                for record in current_batch:
                                    try:
                                        self.env['ddn.property.info'].create(record)
                                        created_records += 1
                                    except Exception as e:
                                        _logger.error(f"Error creating record: {str(e)}")
                                        skipped_records.append({**record, 'error_reason': str(e)}) # Add to skip list
                                
                                _logger.info(f'Batch processed. Total created: {created_records}')
                                current_batch = []
                            except Exception as e:
                                _logger.error(f"Error creating batch: {str(e)}")
                                raise
                    else:
                        # Log as skipped
                        duplicate_field = 'Property ID' if self.import_type == 'property_id' else 'UPIC'
                        duplicate_value = row.get('property_id') if self.import_type == 'property_id' else row.get('upic_no')
                        row['error_reason'] = f'Duplicate {duplicate_field}'
                        skipped_records.append(row)
                        processed_records += 1
                        _logger.info(f"Skipped duplicate {duplicate_field}: {duplicate_value}")

                except Exception as e:
                    _logger.error(f"Error processing row: {str(e)}")
                    continue

            # Process remaining records in the last batch
            if current_batch:
                try:
                    for record in current_batch:
                        try:
                            self.env['ddn.property.info'].create(record)
                            created_records += 1
                        except Exception as e:
                             _logger.error(f"Error creating record: {str(e)}")
                             skipped_records.append({**record, 'error_reason': str(e)})

                except Exception as e:
                    _logger.error(f"Error creating final batch: {str(e)}")
                    raise

            _logger.info(f"Import completed. Total records processed: {processed_records}, Created: {created_records}, Skipped: {len(skipped_records)}")

            # Generate error file if there are skipped records
            if skipped_records:
                error_wb = openpyxl.Workbook()
                error_sheet = error_wb.active
                
                # Write headers
                headers = ['property_id', 'owner_name', 'error_reason']
                for col, header in enumerate(headers, 1):
                    error_sheet.cell(row=1, column=col, value=header)
                
                # Write skipped records
                for row_idx, record in enumerate(skipped_records, 2):
                    error_sheet.cell(row=row_idx, column=1, value=str(record.get('property_id')))
                    error_sheet.cell(row=row_idx, column=2, value=str(record.get('owner_name')))
                    error_sheet.cell(row=row_idx, column=3, value=str(record.get('error_reason')))
                
                # Save error file
                error_file_path = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                error_wb.save(error_file_path.name)
                
                with open(error_file_path.name, 'rb') as f:
                    error_file_content = f.read()
                
                # Update wizard with error file
                self.write({
                    'error_file': base64.b64encode(error_file_content),
                    'error_filename': 'import_errors.xlsx'
                })
                
                # Clean up temporary file
                os.unlink(error_file_path.name)
                
                return {
                    'type': 'ir.actions.act_window',
                    'res_model': 'property.import.wizard',
                    'res_id': self.id,
                    'view_mode': 'form',
                    'target': 'new',
                }

        except UserError as ue:
             raise ue
        except Exception as e:
            _logger.error(f"Import failed: {str(e)}")
            raise ValidationError('Import failed: %s' % str(e))
        
        return {'type': 'ir.actions.act_window_close'}

    def action_test_record(self):
        zone_id = self.env['ddn.zone'].search([('name', '=', 'TEST ZONE')], limit=1)
        ward_id = self.env['ddn.ward'].search([('name', '=', 'TEST WARD')], limit=1)
        colony_id = self.env['ddn.colony'].search([('name', '=', 'TEST COLONY')], limit=1)
        
        if not all([zone_id, ward_id, colony_id]):
            raise ValidationError('TEST records not found in the database.')
        
        self.env['ddn.property.info'].create({
            'zone_id': zone_id.id,
            'ward_id': ward_id.id,
            'colony_id': colony_id.id,
            'upic_no': 'TEST001',
            'property_id': 'TEST001',
            'unit_no': '1'
        })
        
        return {'type': 'ir.actions.act_window_close'} 