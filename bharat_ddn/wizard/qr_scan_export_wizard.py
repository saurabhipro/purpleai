# -*- coding: utf-8 -*-
from odoo import models, fields
from odoo.exceptions import UserError
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os
from datetime import datetime

class QRScanExportWizard(models.TransientModel):
    _name = 'qr.scan.export.wizard'
    _description = 'Export QR Scans to Excel'

    excel_file = fields.Binary('Excel File', readonly=True)
    filename = fields.Char('Filename', readonly=True)
    state = fields.Selection([
        ('choose', 'Choose'),
        ('download', 'Download')
    ], default='choose')

    def action_export_excel(self):
        """Export ALL QR scan records to Excel"""
        # Always export ALL scans, ignore selection
        qr_scans = self.env['ddn.qr.scan'].search([], order='scan_time desc')
        
        if not qr_scans:
            raise UserError('No QR scan records found to export.')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'QR Scan Report'
        
        # Define header style
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Define headers
        headers = [
            'Scan Date & Time',
            'UUID',
            'Property UPIC',
            'Zone',
            'Ward',
            'Colony',
            'Property Type',
            'Latitude',
            'Longitude',
            'Scanned URL'
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, scan in enumerate(qr_scans, start=2):
            property_record = scan.property_id
            
            # Format scan time
            scan_time_str = ''
            if scan.scan_time:
                scan_time_str = scan.scan_time.strftime('%Y-%m-%d %H:%M:%S')
            
            data = [
                scan_time_str,
                scan.uuid or '',
                scan.upic_no or '',
                property_record.zone_id.name if property_record and property_record.zone_id else '',
                property_record.ward_id.name if property_record and property_record.ward_id else '',
                property_record.colony_id.name if property_record and property_record.colony_id else '',
                property_record.property_type.name if property_record and property_record.property_type else '',
                property_record.latitude if property_record else '',
                property_record.longitude if property_record else '',
                scan.scan_url or ''
            ]
            
            for col_idx, value in enumerate(data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Read file content
        with open(temp_file.name, 'rb') as f:
            file_content = f.read()
        
        # Clean up temp file
        os.unlink(temp_file.name)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'qr_scan_report_{timestamp}.xlsx'
        
        # Write to wizard
        self.write({
            'excel_file': base64.b64encode(file_content),
            'filename': filename,
            'state': 'download'
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'qr.scan.export.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
            'context': self.env.context,
        }
    
    def action_close(self):
        """Close the wizard"""
        return {'type': 'ir.actions.act_window_close'}
