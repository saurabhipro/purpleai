# -*- coding: utf-8 -*-
from odoo import models, fields, api
from odoo.exceptions import ValidationError, UserError
import base64
import os
import openpyxl
import tempfile
import logging

_logger = logging.getLogger(__name__)

class DeleteSurveysWizard(models.TransientModel):
    _name = 'delete.surveys.wizard'
    _description = 'Wizard to Delete Surveys from Excel File or Text Input'

    input_method = fields.Selection([
        ('excel', 'Upload Excel File'),
        ('text', 'Paste UPIC Numbers')
    ], string='Input Method', default='text', required=True)
    
    excel_file = fields.Binary('Excel File', help="Upload an Excel file with UPIC numbers")
    filename = fields.Char('Filename')
    
    upic_text = fields.Text('UPIC Numbers', help="Enter comma-separated UPIC numbers (e.g., UPIC001, UPIC002, UPIC003)")
    
    result_file = fields.Binary('Result File', readonly=True)
    result_filename = fields.Char('Result Filename', readonly=True)
    show_result = fields.Boolean('Show Result', default=False)
    
    # Statistics fields
    total_records = fields.Integer('Total Records', readonly=True)
    success_count = fields.Integer('Success Count', readonly=True)
    failed_count = fields.Integer('Failed Count', readonly=True)
    not_found_count = fields.Integer('Not Found Count', readonly=True)

    def action_delete_surveys(self):
        """Process the input (Excel file or text) and delete surveys for the listed UPIC numbers"""
        success_records = []
        failed_records = []
        not_found_records = []
        upic_list = []
        
        try:
            # Get UPIC numbers based on input method
            if self.input_method == 'excel':
                upic_list = self._get_upics_from_excel()
            else:  # text input
                upic_list = self._get_upics_from_text()
            
            if not upic_list:
                raise ValidationError('No UPIC numbers found to process.')
            
            # Process each UPIC
            total_records = len(upic_list)
            for upic_no in upic_list:
                try:
                    # Search for property by UPIC number
                    property_record = self.env['ddn.property.info'].sudo().search([
                        ('upic_no', '=', upic_no)
                    ], limit=1)
                    
                    if not property_record:
                        not_found_records.append({
                            'upic_no': upic_no,
                            'error': 'Property not found'
                        })
                        _logger.warning(f"Property not found for UPIC: {upic_no}")
                        continue
                    
                    # Delete all survey records
                    survey_count = len(property_record.survey_line_ids)
                    if property_record.survey_line_ids:
                        property_record.survey_line_ids.unlink()
                    
                    # Reset property fields to remove survey data
                    # Note: Clearing lat/long will automatically clear the computed digipin field
                    reset_vals = {
                        'property_status': 'pdf_downloaded',
                        'address_line_1': False,
                        'address_line_2': False,
                        'property_id': False,
                        'owner_name': False,
                        'property_type': False,
                        'latitude': False,
                        'longitude': False,
                        'mobile_no': False,
                        'surveyer_id': False,
                    }
                    
                    property_record.write(reset_vals)
                    
                    success_records.append({
                        'upic_no': upic_no,
                        'surveys_deleted': survey_count,
                        'message': 'Success'
                    })
                    
                    _logger.info(f"Successfully deleted surveys for UPIC: {upic_no}")
                    
                except Exception as e:
                    failed_records.append({
                        'upic_no': upic_no,
                        'error': str(e)
                    })
                    _logger.error(f"Error processing UPIC {upic_no}: {str(e)}")
            
            # Generate result Excel file
            result_wb = openpyxl.Workbook()
            
            # Success sheet
            success_sheet = result_wb.active
            success_sheet.title = 'Success'
            success_sheet.append(['UPIC NO', 'Surveys Deleted', 'Status'])
            for record in success_records:
                success_sheet.append([
                    record['upic_no'],
                    record['surveys_deleted'],
                    record['message']
                ])
            
            # Failed sheet
            if failed_records:
                failed_sheet = result_wb.create_sheet('Failed')
                failed_sheet.append(['UPIC NO', 'Error'])
                for record in failed_records:
                    failed_sheet.append([
                        record['upic_no'],
                        record['error']
                    ])
            
            # Not Found sheet
            if not_found_records:
                not_found_sheet = result_wb.create_sheet('Not Found')
                not_found_sheet.append(['UPIC NO', 'Error'])
                for record in not_found_records:
                    not_found_sheet.append([
                        record['upic_no'],
                        record['error']
                    ])
            
            # Save result file
            result_file_path = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            result_wb.save(result_file_path.name)
            
            with open(result_file_path.name, 'rb') as f:
                result_file_content = f.read()
            
            # Update wizard with result file and statistics
            self.write({
                'result_file': base64.b64encode(result_file_content),
                'result_filename': 'delete_surveys_result.xlsx',
                'show_result': True,
                'total_records': total_records,
                'success_count': len(success_records),
                'failed_count': len(failed_records),
                'not_found_count': len(not_found_records),
            })
            
            # Clean up temporary file
            os.unlink(result_file_path.name)
            
            # Return the wizard form with results
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'delete.surveys.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
                'context': self.env.context,
            }
            
        except Exception as e:
            _logger.error(f"Delete surveys failed: {str(e)}")
            raise ValidationError('Operation failed: %s' % str(e))
    
    def action_close(self):
        """Close the wizard"""
        return {'type': 'ir.actions.act_window_close'}
    
    def _get_upics_from_excel(self):
        """Extract UPIC numbers from uploaded Excel file"""
        if not self.excel_file:
            raise ValidationError('Please upload an Excel file.')
        if not self.filename:
            raise ValidationError('Filename is missing.')
        
        ext = os.path.splitext(self.filename)[-1].lower()
        if ext not in ('.xls', '.xlsx'):
            raise ValidationError('Only Excel files (.xls, .xlsx) are supported.')
        
        upic_list = []
        
        # Decode the file
        file_content = base64.b64decode(self.excel_file)
        
        # Read Excel file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file_content)
            tmp.seek(0)
            wb = openpyxl.load_workbook(tmp.name)
            sheet = wb.active
            
            # Get headers
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            
            # Find UPIC column (case-insensitive)
            upic_column_index = None
            for idx, header in enumerate(headers):
                if header and str(header).strip().upper() in ['UPIC', 'UPICNO', 'UPIC_NO', 'UPIC NO']:
                    upic_column_index = idx
                    break
            
            if upic_column_index is None:
                raise ValidationError('UPIC column not found in Excel file. Please ensure the file has a column named "UPIC", "UPICNO", or "UPIC NO".')
            
            # Extract UPIC numbers from each row
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                upic_no = row[upic_column_index]
                
                # Skip empty rows
                if not upic_no:
                    continue
                
                upic_no = str(upic_no).strip()
                if upic_no:
                    upic_list.append(upic_no)
            
            # Clean up temp file
            os.unlink(tmp.name)
        
        return upic_list
    
    def _get_upics_from_text(self):
        """Extract UPIC numbers from comma-separated text input"""
        if not self.upic_text:
            raise ValidationError('Please enter UPIC numbers.')
        
        upic_list = []
        
        # Split by comma, semicolon, newline, or space
        # and clean up each UPIC number
        raw_text = self.upic_text.replace('\n', ',').replace(';', ',').replace('\t', ',')
        upic_numbers = raw_text.split(',')
        
        for upic in upic_numbers:
            upic = upic.strip()
            if upic:  # Only add non-empty values
                upic_list.append(upic)
        
        # Remove duplicates while preserving order
        upic_list = list(dict.fromkeys(upic_list))
        
        return upic_list