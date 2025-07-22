from odoo import models, fields
from datetime import datetime
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pytz
from openpyxl.chart import PieChart, Reference, BarChart
from collections import defaultdict



class DdnReport(models.TransientModel):
    _name = 'ddn.report'
    _description = 'DDN Report Wizard'

    date_from = fields.Date(
        string='Start Date',
        required=True,
        default=lambda self: fields.Date.context_today(self)
    )
    date_to = fields.Date(
        string='End Date',
        required=True,
        default=lambda self: fields.Date.context_today(self)
    )
    company_id = fields.Many2one(
        'res.company',
        string='Company',
        required=True,
        default=lambda self: self.env.company
    )
    zone_id = fields.Many2many('ddn.zone', string='Zone')
    ward_id = fields.Many2many('ddn.ward', string='Ward')


    def generate_xlsx_report(self):
        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Survey Report"

        # Create dashboard worksheet
        ws_dash = wb.create_sheet(title="Dashboard")

        # Create duplicate property ID worksheet
        ws_duplicate = wb.create_sheet(title="Duplicate Property ID")

        # Styles
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")    # Light red
        orange_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold/Orange for visit again
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        
        header_font = Font(bold=True)
        title_font = Font(bold=True, size=25)
        align_left = Alignment(horizontal="left", vertical="center")
        brown_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
        white_header_font = Font(bold=True, color="FFFFFF")
        blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Strong blue
        green_fill_dash = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Forest green
        orange_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")  # Orange
        purple_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # Purple
        section_title_font = Font(bold=True, size=16, color="1F4E78")  # Dark blue
        white_bold_font = Font(bold=True, color="FFFFFF")

        # Title Row
        ws.merge_cells('A2:T2')  # Updated to accommodate new columns
        title_cell = ws['A2']
        title_cell.value = "Survey Report"
        title_cell.font = title_font
        title_cell.fill = header_fill
        title_cell.alignment = align_left
        title_cell.border = border

        # Date Range
        ws.cell(row=4, column=1, value="Date From").border = border
        ws.cell(row=4, column=1).alignment = align_left
        ws.cell(row=4, column=2, value=self.date_from.strftime('%d-%m-%Y') if self.date_from else "").border = border
        ws.cell(row=5, column=1, value="Date To").border = border
        ws.cell(row=5, column=2, value=self.date_to.strftime('%d-%m-%Y') if self.date_to else "").border = border

        # Headers
        headers = [
            "UPIC No", "Property Id", "Zone", "Ward", "Colony", "Owner Name", 
            "Father Name", "Mobile No", "Address Line 1", "Address Line 2", 
            "Latitude", "Longitude", "Total Floors", "Floor Number", 
            "Surveyor", "Surveyor Datetime", "Property Type", "Microsite Url",
            "Property Status", "Is Solar", "Is Rain Water Harvesting", "Old Mobile", "Old Property ID"
        ]

        # Build your domain
        domain = [('company_id', '=', self.company_id.id)]
        
        # Add date filters if provided
        if self.date_from:
            domain.append(('create_date', '>=', self.date_from))
        if self.date_to:
            domain.append(('create_date', '<=', self.date_to))
        
        # Add zone filter if provided
        if self.zone_id:
            domain.append(('property_id.zone_id', 'in', self.zone_id.ids))
        
        # Add ward filter if provided
        if self.ward_id:
            domain.append(('property_id.ward_id', 'in', self.ward_id.ids))

        # Get unique properties that have surveys, not all survey records
        survey_records = self.env['ddn.property.survey'].search(domain)
        unique_property_ids = survey_records.mapped('property_id.id')
        
        # Get the latest survey for each unique property
        records = []
        for property_id in unique_property_ids:
            property_surveys = survey_records.filtered(lambda r: r.property_id.id == property_id)
            # Get the most recent survey for this property
            latest_survey = max(property_surveys, key=lambda r: r.create_date)
            records.append(latest_survey)

        # Now you can use records for summaries, charts, and data table
        # --- Prepare summary data ---
        # 1. Surveyor-wise Count
        surveyor_counts = {}
        for rec in records:
            name = rec.surveyer_id.name or 'Unknown'
            surveyor_counts[name] = surveyor_counts.get(name, 0) + 1

        # 2. Property Status Count
        status_counts = {}
        for rec in records:
            status = rec.property_id.property_status or 'Unknown'
            status_counts[status] = status_counts.get(status, 0) + 1

        # 3. Property Type Count
        type_counts = {}
        for rec in records:
            ptype = rec.property_id.property_type.name or 'Unknown'
            type_counts[ptype] = type_counts.get(ptype, 0) + 1

        # --- Prepare day-wise data ---
        from collections import defaultdict
        day_status_counts = defaultdict(lambda: defaultdict(int))
        all_statuses = set()
        for rec in records:
            day = rec.create_date.strftime('%d-%m-%Y') if rec.create_date else ''
            status = rec.property_id.property_status or 'Unknown'
            day_status_counts[day][status] += 1
            all_statuses.add(status)
        all_statuses = sorted(all_statuses)
        days_sorted = sorted(day_status_counts.keys(), key=lambda d: datetime.strptime(d, '%d-%m-%Y'))

        # --- Write summaries side by side at the top of dashboard sheet ---
        summary_start_row = 3

        # Surveyor Summary
        ws_dash.cell(row=summary_start_row, column=1, value="Surveyor").font = white_bold_font
        ws_dash.cell(row=summary_start_row, column=1).fill = blue_fill
        ws_dash.cell(row=summary_start_row, column=2, value="Count").font = white_bold_font
        ws_dash.cell(row=summary_start_row, column=2).fill = blue_fill
        row1 = summary_start_row + 1
        for name, count in surveyor_counts.items():
            ws_dash.cell(row=row1, column=1, value=name)
            ws_dash.cell(row=row1, column=2, value=count)
            row1 += 1
        ws_dash.cell(row=row1, column=1, value="Grand Total").font = white_bold_font
        ws_dash.cell(row=row1, column=2, value=sum(surveyor_counts.values())).font = white_bold_font

        # Property Status Summary
        ws_dash.cell(row=summary_start_row, column=4, value="Status").font = white_bold_font
        ws_dash.cell(row=summary_start_row, column=4).fill = green_fill_dash
        ws_dash.cell(row=summary_start_row, column=5, value="Count").font = white_bold_font
        ws_dash.cell(row=summary_start_row, column=5).fill = green_fill_dash
        row2 = summary_start_row + 1
        for status, count in status_counts.items():
            ws_dash.cell(row=row2, column=4, value=status)
            ws_dash.cell(row=row2, column=5, value=count)
            row2 += 1
        ws_dash.cell(row=row2, column=4, value="Grand Total").font = white_bold_font
        ws_dash.cell(row=row2, column=5, value=sum(status_counts.values())).font = white_bold_font

        # Property Type Summary
        ws_dash.cell(row=summary_start_row, column=7, value="Type").font = white_bold_font
        ws_dash.cell(row=summary_start_row, column=7).fill = orange_fill
        ws_dash.cell(row=summary_start_row, column=8, value="Count").font = white_bold_font
        ws_dash.cell(row=summary_start_row, column=8).fill = orange_fill
        row3 = summary_start_row + 1
        for ptype, count in type_counts.items():
            ws_dash.cell(row=row3, column=7, value=ptype)
            ws_dash.cell(row=row3, column=8, value=count)
            row3 += 1
        ws_dash.cell(row=row3, column=7, value="Grand Total").font = white_bold_font
        ws_dash.cell(row=row3, column=8, value=sum(type_counts.values())).font = white_bold_font

        # --- Pie Charts ---
        charts_row = max(row1, row2, row3) + 2
        pie1 = PieChart()
        pie1.title = "Surveyor Distribution"
        labels1 = Reference(ws_dash, min_col=1, min_row=summary_start_row+1, max_row=row1-1)
        data1 = Reference(ws_dash, min_col=2, min_row=summary_start_row, max_row=row1-1)
        pie1.add_data(data1, titles_from_data=True)
        pie1.set_categories(labels1)
        pie1.width = 4
        pie1.height = 4

        pie2 = PieChart()
        pie2.title = "Status Distribution"
        labels2 = Reference(ws_dash, min_col=4, min_row=summary_start_row+1, max_row=row2-1)
        data2 = Reference(ws_dash, min_col=5, min_row=summary_start_row, max_row=row2-1)
        pie2.add_data(data2, titles_from_data=True)
        pie2.set_categories(labels2)
        pie2.width = 4
        pie2.height = 4

        pie3 = PieChart()
        pie3.title = "Type Distribution"
        labels3 = Reference(ws_dash, min_col=7, min_row=summary_start_row+1, max_row=row3-1)
        data3 = Reference(ws_dash, min_col=8, min_row=summary_start_row, max_row=row3-1)
        pie3.add_data(data3, titles_from_data=True)
        pie3.set_categories(labels3)
        pie3.width = 4
        pie3.height = 4

        ws_dash.add_chart(pie1, f"A{charts_row}")
        ws_dash.add_chart(pie2, f"D{charts_row}")
        ws_dash.add_chart(pie3, f"G{charts_row}")

        # --- Day-wise Status Table ---
        table_row = charts_row + 10
        ws_dash.cell(row=table_row, column=1, value="Date").font = white_bold_font
        for idx, status in enumerate(all_statuses, 2):
            ws_dash.cell(row=table_row, column=idx, value=status).font = white_bold_font

        for i, day in enumerate(days_sorted, 1):
            ws_dash.cell(row=table_row + i, column=1, value=day)
            for j, status in enumerate(all_statuses, 2):
                ws_dash.cell(row=table_row + i, column=j, value=day_status_counts[day][status])

        # --- Bar Chart for Day-wise Property Status ---
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.title = "Day-wise Property Status"
        bar_chart.style = 13
        bar_chart.y_axis.title = 'Count'
        bar_chart.x_axis.title = 'Date'

        data_ref = Reference(ws_dash, min_col=2, min_row=table_row, max_col=1+len(all_statuses), max_row=table_row+len(days_sorted))
        cats_ref = Reference(ws_dash, min_col=1, min_row=table_row+1, max_row=table_row+len(days_sorted))
        bar_chart.add_data(data_ref, titles_from_data=True)
        bar_chart.set_categories(cats_ref)
        bar_chart.width = 12
        bar_chart.height = 6
        ws_dash.add_chart(bar_chart, f"A{table_row + len(days_sorted) + 2}")

        # --- Data table starts right after the date range ---
        data_start_row = 7  # or 8, right after the date range

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=data_start_row, column=col_num, value=header)
            cell.font = white_header_font
            cell.fill = brown_fill
            cell.alignment = align_left
            cell.border = border

        # Add filter to the header row (only for the data table)
        last_data_row = data_start_row + len(records)
        ws.auto_filter.ref = f"A{data_start_row}:W{last_data_row}"  # Updated to include new columns

        # Freeze panes so header is always visible
        ws.freeze_panes = ws[f"A{data_start_row+1}"]

        # Write data rows
        row = data_start_row + 1
        for rec in records:
            prop = rec.property_id
            
            # Look up old mobile and property ID from property_id_data model
            old_mobile = ''
            old_property_id = ''
            
            if rec.owner_name and (rec.address_line_1 or rec.address_line_2):
                # Create address string for matching
                address_str = f"{rec.address_line_1 or ''} {rec.address_line_2 or ''}".strip()
                
                # Search in property_id_data model based on owner name and address
                property_data_records = self.env['property.id.data'].search([
                    ('owner_name', '=', rec.owner_name),
                    ('address', 'ilike', address_str)
                ], limit=1)
                
                if property_data_records:
                    old_mobile = property_data_records.mobile_no or ''
                    old_property_id = property_data_records.property_id or ''
            
            values = [
                prop.upic_no or '',
                prop.property_id or '',
                prop.zone_id.name or '',
                prop.ward_id.name or '',
                prop.colony_id.name or '',
                rec.owner_name or '',
                rec.father_name or '',
                rec.mobile_no or '',
                rec.address_line_1 or '',
                rec.address_line_2 or '',
                rec.latitude or '',
                rec.longitude or '',
                rec.total_floors or '',
                rec.floor_number or '',
                rec.surveyer_id.name or '',
                rec.create_date.strftime('%d-%m-%Y') if rec.create_date else '',
                prop.property_type.name or '',
                prop.microsite_url or '',
                prop.property_status or '',
                'Yes' if rec.is_solar else 'No',
                'Yes' if rec.is_rainwater_harvesting else 'No',
                old_mobile,
                old_property_id,
            ]
            for col_num, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_num, value=val)
                cell.border = border
                cell.alignment = align_left
                # Apply color coding for Property Status column
                if col_num == 19:  # Property Status column
                    if prop.property_status == 'surveyed':
                        cell.fill = green_fill
                    elif prop.property_status == 'visit_again':
                        cell.fill = orange_fill
                # Apply color coding for solar and rainwater harvesting columns
                elif col_num == 20:  # Is Solar column
                    cell.fill = green_fill if rec.is_solar else red_fill
                elif col_num == 21:  # Is Rain Water Harvesting column
                    cell.fill = green_fill if rec.is_rainwater_harvesting else red_fill
            row += 1

        # --- Set all columns to the width of the "UPIC No" column ---
        upic_col_letter = 'A'
        upic_width = ws.column_dimensions[upic_col_letter].width
        if not upic_width:
            # If not set, calculate based on header length
            upic_width = max(len("UPIC No"), max((len(str(cell.value)) for cell in ws[upic_col_letter] if cell.value), default=8)) + 2

        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = upic_width

        # --- Duplicate Property ID Tab ---
        # Find all duplicate property IDs from property_id_data model
        duplicate_property_ids = self.env['property.id.data'].search([])
        
        # Group by property_id and find duplicates
        property_id_groups = {}
        for record in duplicate_property_ids:
            prop_id = record.property_id
            if prop_id not in property_id_groups:
                property_id_groups[prop_id] = []
            property_id_groups[prop_id].append(record)
        
        # Filter only groups with more than one record (duplicates)
        duplicate_groups = {k: v for k, v in property_id_groups.items() if len(v) > 1}
        
        # Prepare duplicate data for the sheet
        duplicate_data = []
        for prop_id, records in duplicate_groups.items():
            for record in records:
                duplicate_data.append({
                    'property_id': record.property_id,
                    'owner_name': record.owner_name,
                    'address': record.address,
                    'mobile_no': record.mobile_no,
                    'currnet_tax': record.currnet_tax,
                    'total_amount': record.total_amount,
                    'duplicate_count': len(records)
                })
        
        # Write duplicate property ID headers
        duplicate_headers = [
            "Property ID", "Owner Name", "Address", "Mobile No", 
            "Current Tax", "Total Amount", "Duplicate Count"
        ]
        
        duplicate_start_row = 3
        for col_num, header in enumerate(duplicate_headers, 1):
            cell = ws_duplicate.cell(row=duplicate_start_row, column=col_num, value=header)
            cell.font = white_header_font
            cell.fill = brown_fill
            cell.alignment = align_left
            cell.border = border
        
        # Add filter to the duplicate header row
        last_duplicate_row = duplicate_start_row + len(duplicate_data)
        ws_duplicate.auto_filter.ref = f"A{duplicate_start_row}:G{last_duplicate_row}"
        
        # Freeze panes for duplicate sheet
        ws_duplicate.freeze_panes = ws_duplicate[f"A{duplicate_start_row+1}"]
        
        # Write duplicate data rows
        duplicate_row = duplicate_start_row + 1
        for data in duplicate_data:
            values = [
                data['property_id'],
                data['owner_name'],
                data['address'],
                data['mobile_no'],
                data['currnet_tax'],
                data['total_amount'],
                data['duplicate_count']
            ]
            for col_num, val in enumerate(values, 1):
                cell = ws_duplicate.cell(row=duplicate_row, column=col_num, value=val)
                cell.border = border
                cell.alignment = align_left
                # Highlight duplicate count column
                if col_num == 7:  # Duplicate Count column
                    if data['duplicate_count'] > 2:
                        cell.fill = red_fill
                    else:
                        cell.fill = orange_fill
            duplicate_row += 1
        
        # Set column widths for duplicate sheet
        for col in ws_duplicate.columns:
            col_letter = col[0].column_letter
            ws_duplicate.column_dimensions[col_letter].width = 15

        # Prepare output for download
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_base64 = base64.b64encode(output.read()).decode('utf-8')

        # Attachment
        calcuttaTz = pytz.timezone("Asia/Kolkata")
        filename = f"Survey_Report_{datetime.now(calcuttaTz).strftime('%d-%m-%Y_%H:%M:%S')}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': file_base64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }